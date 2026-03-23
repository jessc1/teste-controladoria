from pathlib import Path
import pandas as pd
import numpy as np
import os
import re
from datetime import datetime
from typing import Tuple, Dict, Optional, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class Transaction:
    base_dir = Path(__file__).resolve().parent
    data_dir = base_dir / "data"
    transactions_dir = data_dir / "transactions"
    PRICES_FILE = data_dir / "prices.xlsx"
    out_dir = base_dir / "out"
    
    def __init__(self):
        self.transactions = []
        self.invalid_rows = []
        self.prices = None
    

    SIDE_OPTIONS = {
        'buy': 'BUY',
        'sell': 'SELL',
        'compra': 'BUY',
        'venda': 'SELL',
    }
    
    def validate_client_document(self, document: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        if document is None or pd.isna(document) or str(document).strip() == '':
            return None, None, None
    
        document = str(document).strip().upper()
        document = re.sub(r'(CPF|CNPJ|\s+)', '', document)
        document= re.sub(r'[^\d\.\-/]' , '', document)
        digits = re.sub(r'\D', '', document)

        if len(digits) == 11:
            document_type = 'CPF'
            formatted = f"{digits[0:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:11]}"
            return formatted, digits, document_type
        elif len(digits) == 14:
            document_type = 'CNPJ'
            formatted = f"{digits[0:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"
            return formatted, digits, document_type
        return None, None, None

    def validate_number(self, value: str) -> Optional[float]:
        if value is None or pd.isna(value):
            return None
        value = str(value).strip()
        try:
            return float(value)
        except (ValueError, TypeError):
            pass
        comma_count = value.count(',')
        dot_count = value.count('.')
        value = value.replace(' ', '')
        if comma_count == 0 and dot_count == 0:
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        
        if dot_count == 0 and comma_count == 1:
            try:
                return float(value.replace(',', '.'))
            except (ValueError, TypeError):
                return None
        
        if comma_count == 0 and dot_count >= 1:
            if dot_count > 1:
                parts = value.rsplit('.', 1)
                value = parts[0].replace('.', '') + '.' + parts[1]
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        
        if comma_count > 0 and dot_count > 0:
            last_comma_pos = value.rfind(',')
            last_dot_pos = value.rfind('.')
            
            if last_comma_pos > last_dot_pos:
                value = value.replace('.', '').replace(',', '.')
            else:
                value = value.replace(',', '')
            
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        
        return None

    def validate_date(self, value: str) -> Optional[str]:
        if value is None or pd.isna(value) or str(value).strip() == '':
            return None
        value = str(value).strip()
        dates = [
            '%d/%m%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%d-%m-%y',
            '%Y-%m-%d',
        ]

        for fmt in dates:
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        return None 

    def validate_side(self, value: str) -> Optional[str]:
        if value is None or pd.isna(value) or str(value).strip() == '':
            return None
        value = str(value).strip().lower()
        return self.SIDE_OPTIONS.get(value)

    def load_prices(self) -> bool:
        try:       
            df = pd.read_excel(self.PRICES_FILE)
            df.columns = df.columns.str.lower().str.strip()
            df['date'] = df['date'].apply(self.validate_date)
            df['price'] = df['price'].apply(self.validate_number)
            df = df.dropna(subset=['date', 'ticker', 'price'])
            self.prices = df
            return True
        except Exception as e:
            print(f"Error in load prices: {e}")
            return False

    def get_price(self, date: str, ticker: str) -> Optional[float]:
        if self.prices is None or date is None or ticker is None:
            return None
        try:
            result = self.prices[(self.prices['date'] == date) & (self.prices['ticker'] == ticker)]['price']
            if len(result) > 0:
                return result.iloc[0]
        except Exception as e:
            print(f"Error in get_price: {e}")
        return None

    def validate_row(self, row: Dict[str, Any], source_file: str) -> Tuple[Optional[Dict], Optional[Dict]]:       
        original_row = row.copy()
        original_row['source_file'] = source_file
        
        try:
            date = self.validate_date(row.get('date', ''))
            if date is None:
                return None, {**original_row, 'invalid_reason': 'invalid_date'}
            
            side = self.validate_side(row.get('side', ''))
            if side is None:
                return None, {**original_row, 'invalid_reason': 'invalid_side'}
            
            client_doc, doc_clean, doc_type = self.validate_and_normalize_document(
                row.get('client_document', '')
            )
            if client_doc is None:
                return None, {**original_row, 'invalid_reason': 'invalid_document'}
            
            quantity = self.validate_number(row.get('quantity', ''))
            broker_fee = self.validate_number(row.get('broker_fee', ''))
            tax = self.validate_number(row.get('tax', ''))
            currency = row.get('currency', '')
            
            if quantity is None or broker_fee is None or tax is None:
                return None, {**original_row, 'invalid_reason': 'invalid_number'}
            
            if quantity <= 0:
                return None, {**original_row, 'invalid_reason': 'invalid_quantity'}
            
            if broker_fee < 0 or tax < 0:
                return None, {**original_row, 'invalid_reason': 'invalid_costs'}
            
            price = self.get_price(date, row.get('ticker', ''))
            if price is None:
                return None, {**original_row, 'invalid_reason': 'invalid_number'}  # Price not found
            
            if price < 0:
                return None, {**original_row, 'invalid_reason': 'invalid_price'}
            
            if side == 'BUY':
                gross_amount = quantity * price
            else:  
                gross_amount = -(quantity * price)
            
            total_costs = broker_fee + tax
            net_amount = gross_amount - total_costs
            
            cleaned_row = {
                'trade_id': row.get('trade_id', ''),
                'account_id': row.get('account_id', ''),
                'client_document': client_doc,
                'document_clean': doc_clean,
                'document_type': doc_type,
                'date': date,
                'ticker': row.get('ticker', ''),
                'side': side,
                'quantity': quantity,
                'price': price,
                'broker_fee': broker_fee,
                'tax': tax,
                'gross_amount': gross_amount,
                'total_costs': total_costs,
                'net_amount': net_amount,
                'currency': currency,
                'source_file': source_file,
            }
            
            return cleaned_row, None
            
        except Exception as e:
            return None, {**original_row, 'invalid_reason': 'invalid_number'}
    
    def read_transactions_files(self) -> pd.DataFrame:
        all_data = []
        
        if not os.path.exists(self.transactions_dir):
            print(f"Directory not found: {self.transactions_dir}")
            return pd.DataFrame()
        
        for filename in sorted(os.listdir(self.transactions_dir)):
            if not filename.endswith('.csv'):
                continue
            
            filepath = os.path.join(self.transactions_dir, filename)
            
            try:
                encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'ascii']
                df = None
                
                for encoding in encodings:
                    try:
                        df = pd.read_csv(filepath, encoding=encoding, sep=',')
                        break
                    except UnicodeDecodeError:
                        continue
                
                if df is None:
                    print(f"Could not read {filename} with any encoding")
                    continue
                
                # Normalize column names
                df.columns = df.columns.str.lower().str.strip()
                
                # Add source file
                df['_source_file'] = filename
                
                all_data.append(df)
                print(f"Loaded {filename} with {len(df)} rows")
                
            except Exception as e:
                print(f"Error reading {filename}: {e}")
                continue
        
        if not all_data:
            return pd.DataFrame()
        
        return pd.concat(all_data, ignore_index=True)
    


    def process(self) -> bool:            
        print("Start to process...")
        os.makedirs(self.out_dir, exist_ok=True)
        print("Loading prices...")
        if not self.load_prices():
                print("Error in load prices")
                return False               
        print("Reading")
        transactions_df =  self.read_transactions_files()
        if transactions_df.empty:
            print("transactions is empty")
            return False
        print(f"Total read: {len(transactions_df)}")
        print("Processing transactions...")
        for idx, row in transactions_df.iterrows():
            source = row.pop('_source_file') if '_source_file' in row else 'unknown'
            cleaned_row, error_row = self.validate_row(row.to_dict(), source)
            print('cleaned_row', cleaned_row, 'error_row', error_row)
            if cleaned_row:
                self.transactions.append(cleaned_row)
            else:
                self.invalid_rows.append(error_row)
            
            print(f"Transactions: {len(self.transactions)}")
            print(f"Invalid rows: {len(self.invalid_rows)}")
            
            print("Handling duplicates...")
            self.transactions = self.handle_duplicates(self.transactions)
            print(f"Deduplications: {len(self.transactions)}")
            
            print("Output...")
            self.generate_clean_transactions()
            self.generate_invalid_rows()
            self.generate_daily_positions()
            
            print("complete!")
            return True

    def handle_duplicates(self, transactions: list) -> list:
        
        if not transactions:
            return transactions
        
        df = pd.DataFrame(transactions)       
        df['batch_number'] = df['source_file'].str.extract(r'(\d+)').astype(int)
        df = df.sort_values('batch_number', ascending=False)        
        df = df.drop_duplicates(subset=['trade_id'], keep='first')        
        df = df.drop('batch_number', axis=1)
        return df.to_dict('records')
        
    def generate_clean_transactions(self):
        if not self.transactions:
            df = pd.DataFrame()
        else:
            df = pd.DataFrame(self.transactions)
        
        output_path = os.path.join(self.out_dir, 'clean_transactions.csv')        
        columns = [
            'trade_id', 'account_id', 'client_document', 'document_clean', 
            'document_type', 'date', 'ticker', 'side', 'quantity', 'price',
            'broker_fee', 'tax', 'gross_amount', 'total_costs', 'net_amount',
            'currency', 'source_file'
        ]
        
        if not df.empty:
            df = df[columns]
            numeric_cols = ['quantity', 'price', 'broker_fee', 'tax', 'gross_amount', 'total_costs', 'net_amount']
            for col in numeric_cols:
                df[col] = df[col].astype(float)
        else:
            df = pd.DataFrame(columns=columns)
        
        df.to_csv(output_path, index=False, encoding='utf-8')
        print(f"Generated {output_path}")

    def generate_invalid_rows(self):
        if not self.invalid_rows:
            output_path = os.path.join(self.out_dir, 'invalid_rows.csv')
            df = pd.DataFrame()
            df.to_csv(output_path, index=False, encoding='utf-8')
            print(f"Generated {output_path} (empty)")
            return
        else:
            df = pd.DataFrame(self.invalid_rows)
            cols = [c for c in df.columns if c not in ['invalid_reason', 'source_file']]
            cols.extend(['invalid_reason', 'source_file'])
            df = df[[c for c in cols if c in df.columns]]
            output_path = os.path.join(self.out_dir, 'invalid_rows.csv')
            df.to_csv(output_path, index=False, encoding='utf-8')
            print(f"Generated {output_path}")
    
    def format_currency(self, value: float) -> str:
        if value is None or pd.isna(value):
            return "R$ 0.00"
        val = abs(value)
        formatted = f"R$ {val:,.2f}".replace(',', '#').replace('.', ',').replace('#', '.')
        return formatted

    def generate_daily_positions(self):
        if not self.transactions:
            output_path = os.path.join(self.out_dir, 'daily_positions.xlsx')
            df = pd.DataFrame()
            df.to_excel(output_path, index=False)
            print(f"Generated {output_path} (empty)")
            return
            
        df = pd.DataFrame(self.transactions)
        grouped = df.groupby(['date', 'ticker']).agg({
            'gross_amount': 'sum',
            'quantity': 'sum',
            'price': lambda x: np.average(x, weights=df.loc[x.index, 'quantity']),
            'total_costs': 'sum',
        }).reset_index()
            
        grouped.columns = ['date', 'ticker', 'gross_amount', 'total_quantity', 'avg_trade_price', 'total_costs']
        grouped['date'] = pd.to_datetime(grouped['date']).dt.strftime('%d/%m/%Y')
        grouped['gross_amount_formatted'] = grouped['gross_amount'].apply(self.format_currency)
        grouped['avg_trade_price_formatted'] = grouped['avg_trade_price'].apply(self.format_currency)
        grouped['total_costs_formatted'] = grouped['total_costs'].apply(self.format_currency)
        output = grouped[['date', 'ticker', 'gross_amount_formatted', 'avg_trade_price_formatted', 'total_costs_formatted']].copy()
        output.columns = ['date', 'ticker', 'gross_amount', 'avg_trade_price', 'total_costs']
        output_path = os.path.join(self.out_dir, 'daily_positions.xlsx')
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            output.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 10
            worksheet.column_dimensions['C'].width = 16
            worksheet.column_dimensions['D'].width = 16
            worksheet.column_dimensions['E'].width = 16
            for row in worksheet.iter_rows(min_row=2, max_row=len(output_df)+1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
        print(f"Generated {output_path}")
    
